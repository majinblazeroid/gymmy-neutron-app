#!/usr/bin/env python3
"""
Migration script: Gym Ting.xlsx → Supabase

Usage:
    cd migration
    source venv/bin/activate

    # Just load exercises + warmups into DB (do this first):
    python migrate.py --seed-only

    # Import historical workout data from spreadsheet:
    python migrate.py --file "/Users/nirvikgill/Downloads/Gym Ting.xlsx"

    # Test parser without writing to DB:
    python migrate.py --file "/Users/nirvikgill/Downloads/Gym Ting.xlsx" --dry-run
"""

import argparse
import re
import os
import sys
import datetime
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from supabase import create_client, Client

# ─────────────────────────────────────────────────────────
# Env / Supabase setup
# ─────────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")
load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL") or os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")

# ─────────────────────────────────────────────────────────
# Seed data
# ─────────────────────────────────────────────────────────

EXERCISE_DEFS = [
    dict(name="Back Extensions",       type="bodyweight",  suggested_sets=1, suggested_reps="20 reps",      notes="Supports timed hold, reps, or weighted reps"),
    dict(name="Side Hyper Extensions", type="unilateral",  suggested_sets=2, suggested_reps="AMRAP",         notes="Left side first. Match reps on right."),
    dict(name="RDL",                   type="weighted",    suggested_sets=3, suggested_reps="6-8",           default_unit="kg"),
    dict(name="Pull Ups",              type="bodyweight",  suggested_sets=3, suggested_reps="AMRAP",         notes="Optional weight for weighted pull-ups."),
    dict(name="Bench Press",           type="weighted",    suggested_sets=3, suggested_reps="6-8",           default_unit="kg"),
    dict(name="Front Rack Carry",      type="timed_carry", suggested_sets=3, suggested_reps="40s total (10s/leg)", default_unit="lbs", notes="Stationary hold. 10s per leg, 40s total per set."),
    dict(name="Barbell Squat",         type="weighted",    suggested_sets=3, suggested_reps="8-10",          default_unit="kg"),
    dict(name="Chest Supported Row",   type="weighted",    suggested_sets=3, suggested_reps="8-10",          default_unit="lbs"),
    dict(name="Dead Bug",              type="bodyweight",  suggested_sets=3, suggested_reps="8 per side",    notes="Optional weight."),
    dict(name="OHP",                   type="weighted",    suggested_sets=3, suggested_reps="5-6",           default_unit="kg"),
    dict(name="Incline DB Bench",      type="weighted",    suggested_sets=3, suggested_reps="4-6",           default_unit="kg", notes="Optional — only done some weeks."),
]

DAY_A_EXERCISES = ["Back Extensions", "Side Hyper Extensions", "RDL", "Pull Ups", "Bench Press", "Front Rack Carry"]
DAY_B_EXERCISES = ["Back Extensions", "Side Hyper Extensions", "Barbell Squat", "Chest Supported Row", "Dead Bug", "OHP", "Incline DB Bench"]

WARMUPS_A = [
    dict(day="A", name="Hamstring Sweeps",      prescription="10-12 a side",                   order=1),
    dict(day="A", name="Glute Bridges",         prescription="10-12 reps, 1s pause at top",    order=2),
    dict(day="A", name="Lateral Lunges",        prescription="6-8 a side",                     order=3),
    dict(day="A", name="Lat Stretch",           prescription="20-30s hold",                    order=4),
    dict(day="A", name="Light RDL",             prescription="8 reps light + 5 reps at 50%",   order=5),
]
WARMUPS_B = [
    dict(day="B", name="Knee Over Toe Rocks",   prescription="12-15 a side",                   order=1),
    dict(day="B", name="Ankle Pulses",          prescription="10-15 a side",                   order=2),
    dict(day="B", name="90-90 Hips",            prescription="6-8 a side",                     order=3),
    dict(day="B", name="Couch Stretch",         prescription="20-30s a side",                  order=4),
    dict(day="B", name="Light Abductor Machine",prescription="2-3 sets, find height",           order=5),
    dict(day="B", name="Bodyweight Squats",     prescription="8-10 reps",                      order=6),
]

# Spreadsheet row name → DB exercise name
EXERCISE_NAME_MAP = {
    "back extension holds":                    "Back Extensions",
    "hyperextension":                          "Side Hyper Extensions",
    "rdl":                                     "RDL",
    "pull ups":                                "Pull Ups",
    "bench press":                             "Bench Press",
    "front rack carry stationary , one leg 5 s ": "Front Rack Carry",
    "front rack carry":                        "Front Rack Carry",
    "barbell squat (kg)":                      "Barbell Squat",
    "barbell squat":                           "Barbell Squat",
    "ches supported row":                      "Chest Supported Row",
    "chest supported row":                     "Chest Supported Row",
    "dead bug":                                "Dead Bug",
    "dead bug ":                               "Dead Bug",
    "ohp":                                     "OHP",
    "incline db bench (optional)":             "Incline DB Bench",
    "incline db bench":                        "Incline DB Bench",
    "db bench":                                "Incline DB Bench",
}

# ─────────────────────────────────────────────────────────
# Cell parser
# ─────────────────────────────────────────────────────────

def parse_time_obj(t) -> int:
    """datetime.time(1, 40) → openpyxl stores 1:40 as a time; in exercise context = 1min40s = 100s."""
    return t.hour * 60 + t.minute + t.second

def parse_duration_str(s: str):
    """'2 mins' → 120, '1:30' → 90, '90s' → 90, '40 s' → 40. Returns None if not a duration."""
    s = s.strip()
    # mm:ss format
    m = re.match(r"^(\d+):(\d+)", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    # X min(s)
    m = re.match(r"^(\d+\.?\d*)\s*min", s, re.IGNORECASE)
    if m:
        return int(float(m.group(1)) * 60)
    # X s / X sec
    m = re.match(r"^(\d+\.?\d*)\s*s(?:ec)?", s, re.IGNORECASE)
    if m:
        return int(float(m.group(1)))
    return None

def parse_cell(raw, is_carry: bool = False) -> list[dict]:
    """
    Convert a spreadsheet cell value to a list of set dicts.
    Each dict may have: weight, reps, duration_seconds, is_warmup, side, note
    """
    if raw is None:
        return []

    # datetime.time object (openpyxl parsed a time-formatted cell)
    if isinstance(raw, datetime.time):
        return [{"duration_seconds": parse_time_obj(raw)}]

    text = str(raw).strip()
    if not text:
        return []

    # Skip / note only entries
    skip_words = ["skip", "rest", "deload", "holiday", "sick", "cus im tired", "optional"]
    tl = text.lower()
    if any(w in tl for w in skip_words):
        return [{"note": text}]

    sets = []

    # ── Strip common prefixes ──────────────────────────────
    # "Body Weight*20 reps" → "*20 reps"
    # "Body Weight 16 reps" → "16 reps"
    # "no weight*8" → "*8"
    is_bodyweight_prefix = bool(re.match(r"body\s*weight", text, re.IGNORECASE))
    is_no_weight = bool(re.match(r"no\s*weight", text, re.IGNORECASE))
    text_clean = re.sub(r"(?:body\s*weight|no\s*weight)\*?", "", text, flags=re.IGNORECASE).strip()

    # If after stripping it's just "Body Weight" with nothing else
    if not text_clean and (is_bodyweight_prefix or is_no_weight):
        return [{"reps": None, "note": "bodyweight"}]

    # Check for "*20", "20 reps", or just "20" after stripping bodyweight prefix
    if is_bodyweight_prefix or is_no_weight:
        reps_only = re.match(r"^\*?(\d+)\s*reps?(.*)$", text_clean, re.IGNORECASE)
        if reps_only:
            note = reps_only.group(2).strip().strip("()")
            return [{"reps": int(reps_only.group(1)), "note": note or None}]

    # ── Timed: whole cell is a duration ───────────────────
    dur = parse_duration_str(text_clean if (is_bodyweight_prefix or is_no_weight) else text)
    if dur is not None:
        # extract optional note in parens
        note_m = re.search(r"\(([^)]+)\)", text)
        note = note_m.group(1).strip() if note_m else None
        return [{"duration_seconds": dur, "note": note}]

    # ── "X a side" / unilateral ───────────────────────────
    side_m = re.match(r"(\d+)\*?\s*a\s*side", text, re.IGNORECASE)
    if side_m:
        reps = int(side_m.group(1))
        sets.append({"reps": reps, "side": "left"})
        sets.append({"reps": reps, "side": "right"})
        return sets

    # ── "7* a side (left first) + 9 a side (right first)" ─
    complex_side = re.findall(r"(\d+)\*?\s*a\s*side", text, re.IGNORECASE)
    if complex_side and len(complex_side) >= 2:
        for r in complex_side:
            sets.append({"reps": int(r), "side": "left"})
            sets.append({"reps": int(r), "side": "right"})
        return sets

    # ── "3 sets alternate X" ──────────────────────────────
    alt_m = re.match(r"(\d+)\s*sets?\s*(?:alternate\s*)?(\d+)", text, re.IGNORECASE)
    if alt_m:
        n_sets = int(alt_m.group(1))
        reps = int(alt_m.group(2))
        return [{"reps": reps} for _ in range(n_sets)]

    # ── Multi-set parsing ─────────────────────────────────
    # Split on "+" but not inside parentheses
    def split_sets(s: str) -> list[str]:
        parts = []
        depth = 0
        current = []
        for ch in s:
            if ch == "(":
                depth += 1
                current.append(ch)
            elif ch == ")":
                depth -= 1
                current.append(ch)
            elif ch == "+" and depth == 0:
                parts.append("".join(current).strip())
                current = []
            else:
                current.append(ch)
        if current:
            parts.append("".join(current).strip())
        return [p for p in parts if p]

    # Also split on spaces between complete tokens (e.g. "70*7 80*3")
    def split_space_sets(s: str) -> list[str]:
        return re.findall(r"[\w.*]+(?:\([^)]*\))?", s)

    raw_parts = split_sets(text)

    # If single part, try splitting by space for patterns like "70*7 80*3"
    if len(raw_parts) == 1:
        space_parts = split_space_sets(raw_parts[0])
        if len(space_parts) > 1 and all(re.search(r"\d+\*\d+", p) for p in space_parts):
            raw_parts = space_parts

    last_weight = None

    for part in raw_parts:
        part = part.strip()
        if not part:
            continue

        is_warmup = bool(re.search(r"\(warmup\)", part, re.IGNORECASE))
        part_clean = re.sub(r"\(warmup\)", "", part, flags=re.IGNORECASE).strip()

        # Extract trailing note in parens
        note_m = re.search(r"\(([^)]*)\)$", part_clean)
        note = None
        if note_m:
            candidate = note_m.group(1).strip()
            # Only treat as note if it doesn't look like a number
            if not re.match(r"^\d+\.?\d*$", candidate):
                note = candidate
            part_clean = part_clean[:note_m.start()].strip()

        # "bar" or "bar * 6" or "bar*6"
        bar_m = re.match(r"^bar\s*\*?\s*(\d+)?$", part_clean, re.IGNORECASE)
        if bar_m:
            reps = int(bar_m.group(1)) if bar_m.group(1) else None
            sets.append({"weight": 0, "reps": reps, "is_warmup": True, "note": "bar"})
            last_weight = 0
            continue

        # "Xkg warmup * Y" or "Xkg warmup"
        kg_warmup_m = re.match(r"^(\d+\.?\d*)\s*kg\s+warmup\s*\*?\s*(\d+)?$", part_clean, re.IGNORECASE)
        if kg_warmup_m:
            weight = float(kg_warmup_m.group(1))
            reps = int(kg_warmup_m.group(2)) if kg_warmup_m.group(2) else None
            sets.append({"weight": weight, "reps": reps, "is_warmup": True, "note": note})
            last_weight = weight
            continue

        # "weight * value [s]" pattern
        wr_m = re.match(r"^(\d+\.?\d*)\s*(?:kg|lbs?)?\s*\*\s*(\d+\.?\d*)\s*(s|sec|min)?", part_clean, re.IGNORECASE)
        if wr_m:
            weight = float(wr_m.group(1))
            value = float(wr_m.group(2))
            unit_suffix = wr_m.group(3)
            last_weight = weight
            if is_carry or unit_suffix in ("s", "sec"):
                secs = int(value * 60) if unit_suffix == "min" else int(value)
                sets.append({"weight": weight, "duration_seconds": secs, "is_warmup": is_warmup, "note": note})
            else:
                sets.append({"weight": weight, "reps": int(value), "is_warmup": is_warmup, "note": note})
            continue

        # Plain number: reps only (reuse last weight if available)
        plain_m = re.match(r"^(\d+)$", part_clean)
        if plain_m:
            reps = int(plain_m.group(1))
            entry: dict = {"reps": reps, "is_warmup": is_warmup}
            if last_weight is not None:
                entry["weight"] = last_weight
            if note:
                entry["note"] = note
            sets.append(entry)
            continue

        # Duration string in a part
        dur = parse_duration_str(part_clean)
        if dur is not None:
            sets.append({"duration_seconds": dur, "is_warmup": is_warmup, "note": note})
            continue

        # Fallback: store as note
        if part_clean:
            sets.append({"note": part_clean})

    if not sets:
        return [{"note": text}]

    return sets

# ─────────────────────────────────────────────────────────
# Seed functions
# ─────────────────────────────────────────────────────────

def seed_exercises(sb: Client) -> dict[str, str]:
    print("Seeding exercises...")
    name_to_id: dict[str, str] = {}
    existing = sb.table("exercises").select("id, name").execute()
    for row in existing.data or []:
        name_to_id[row["name"]] = row["id"]

    to_insert = [ex for ex in EXERCISE_DEFS if ex["name"] not in name_to_id]
    for ex in to_insert:
        res = sb.table("exercises").insert(ex).execute()
        if res.data:
            name_to_id[ex["name"]] = res.data[0]["id"]

    print(f"  {len(name_to_id)} exercises ready ({len(to_insert)} newly inserted)")
    return name_to_id


def seed_templates(sb: Client, name_to_id: dict[str, str]):
    print("Seeding workout templates...")
    sb.table("workout_templates").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    rows = []
    for i, name in enumerate(DAY_A_EXERCISES):
        if name in name_to_id:
            rows.append(dict(day="A", exercise_id=name_to_id[name], order=i + 1))
    for i, name in enumerate(DAY_B_EXERCISES):
        if name in name_to_id:
            rows.append(dict(day="B", exercise_id=name_to_id[name], order=i + 1))
    sb.table("workout_templates").insert(rows).execute()
    print(f"  {len(rows)} template rows inserted")


def seed_warmups(sb: Client):
    print("Seeding warmups...")
    sb.table("warmups").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    sb.table("warmups").insert(WARMUPS_A + WARMUPS_B).execute()
    print(f"  {len(WARMUPS_A + WARMUPS_B)} warmup rows inserted")

# ─────────────────────────────────────────────────────────
# Excel migration
# ─────────────────────────────────────────────────────────

# Spreadsheet layout (Phase 1, "Gym Sched" sheet):
#   Row 1:  date headers in cols C-N
#   Row 3:  "DAY A"
#   Rows 4-9:  Day A exercises (col A = name, col B = prescription, cols C+ = data)
#   Row 12: "DAY B"
#   Rows 13-20: Day B exercises
#   Rows 21+: blank / Phase 2 (skip)

PHASE1_DATE_ROW   = 1   # 1-indexed
PHASE1_DAY_A_ROW  = 3
PHASE1_EX_A_ROWS  = list(range(4, 10))   # rows 4-9  (6 exercises)
PHASE1_DAY_B_ROW  = 12
PHASE1_EX_B_ROWS  = list(range(13, 21))  # rows 13-20 (8 rows, some skipped)
DATA_START_COL    = 3   # col C = index 3 (1-indexed)


def resolve_exercise(raw_name: str) -> str | None:
    """Map a spreadsheet exercise name to the DB name."""
    key = raw_name.strip().lower()
    # Direct lookup
    if key in EXERCISE_NAME_MAP:
        return EXERCISE_NAME_MAP[key]
    # Fuzzy: check if any map key is contained in the cell name or vice versa
    for map_key, db_name in EXERCISE_NAME_MAP.items():
        if map_key.strip() in key or key in map_key.strip():
            return db_name
    return None


def migrate_xlsx(xlsx_path: str, sb: Client, name_to_id: dict[str, str], dry_run: bool):
    print(f"\nOpening {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Gym Sched"]

    # Read all values into a 2D list — 1-indexed for both rows and cols.
    # data[1] = row 1, data[r][1] = col A, data[r][3] = col C, etc.
    data = [[None]]  # dummy row 0
    for row in ws.iter_rows(values_only=True):
        data.append([None] + list(row))

    # Get Phase 1 dates (row 1, cols C onwards)
    dates = []
    for col in range(DATA_START_COL, ws.max_column + 1):
        val = data[PHASE1_DATE_ROW][col]
        if isinstance(val, (datetime.datetime, datetime.date)):
            dates.append((col, val.strftime("%Y-%m-%d")))
        else:
            break  # stop at first non-date

    print(f"Found {len(dates)} weeks: {dates[0][1]} → {dates[-1][1]}")

    # Build exercise row maps
    def build_ex_rows(ex_rows):
        """Return list of (db_exercise_name, row_index) for a day's exercises."""
        result = []
        for row in ex_rows:
            raw = data[row][1]  # col A (1-indexed: col 1)
            if not raw:
                continue
            db_name = resolve_exercise(str(raw))
            if db_name:
                result.append((db_name, row))
            else:
                print(f"  [skip] Unrecognised exercise: '{raw}'")
        return result

    day_a_exercises = build_ex_rows(PHASE1_EX_A_ROWS)
    day_b_exercises = build_ex_rows(PHASE1_EX_B_ROWS)

    sessions_created = 0
    sets_created = 0

    def is_real(val) -> bool:
        """True if a cell has actual logged data (not null or bare 'Body Weight' placeholder)."""
        if val is None:
            return False
        if isinstance(val, datetime.time):
            return True
        s = str(val).strip()
        if not s:
            return False
        # Bare "Body Weight" or "Body Weight " with nothing else = placeholder, skip
        if re.match(r'^body\s*weight\s*$', s, re.IGNORECASE):
            return False
        return True

    for col, date_str in dates:
        for day, ex_list in [("A", day_a_exercises), ("B", day_b_exercises)]:
            # Only create a session if at least one exercise has real logged data
            has_data = any(is_real(data[row][col]) for _, row in ex_list)
            if not has_data:
                print(f"  Day {day} — {date_str}: SKIP (no real data)")
                continue

            print(f"  Day {day} — {date_str}")

            if not dry_run:
                session_res = sb.table("workout_sessions").insert(dict(
                    date=date_str,
                    day=day,
                    pre_feeling=3,
                    post_feeling=3,
                    warmup_completed=False,
                )).execute()
                if not session_res.data:
                    print(f"    ! Failed to create session")
                    continue
                session_id = session_res.data[0]["id"]
            else:
                session_id = "dry-run"

            for db_name, row in ex_list:
                cell_val = data[row][col]
                if cell_val is None:
                    continue

                ex_id = name_to_id.get(db_name)
                if not ex_id:
                    print(f"    ! No ID for {db_name}")
                    continue

                is_carry = db_name == "Front Rack Carry"
                parsed = parse_cell(cell_val, is_carry=is_carry)

                if not parsed:
                    continue

                for j, s in enumerate(parsed):
                    row_data = dict(
                        session_id=session_id,
                        exercise_id=ex_id,
                        set_number=j + 1,
                        weight=s.get("weight"),
                        reps=s.get("reps"),
                        duration_seconds=s.get("duration_seconds"),
                        is_warmup=s.get("is_warmup", False),
                        side=s.get("side"),
                        note=s.get("note"),
                    )
                    if dry_run:
                        print(f"    {db_name} set {j+1}: {s}")
                    else:
                        sb.table("workout_sets").insert(row_data).execute()
                    sets_created += 1

            sessions_created += 1

    print(f"\nDone! {sessions_created} sessions, {sets_created} sets {'(dry run — nothing written)' if dry_run else 'inserted'}.")


# ─────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Gym Ting.xlsx to Supabase")
    parser.add_argument("--file",      required=False, help="Path to Gym Ting.xlsx")
    parser.add_argument("--seed-only", action="store_true", help="Only seed exercises/warmups")
    parser.add_argument("--dry-run",   action="store_true", help="Parse only, don't write to DB")
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_KEY or "your-project" in (SUPABASE_URL or ""):
        print("ERROR: Set NEXT_PUBLIC_SUPABASE_URL and NEXT_PUBLIC_SUPABASE_ANON_KEY in .env.local")
        sys.exit(1)

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    name_to_id = seed_exercises(sb)
    seed_templates(sb, name_to_id)
    seed_warmups(sb)

    if not args.seed_only:
        if not args.file:
            print("ERROR: Provide --file path or use --seed-only")
            sys.exit(1)
        if not Path(args.file).exists():
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        migrate_xlsx(args.file, sb, name_to_id, dry_run=args.dry_run)